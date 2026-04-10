import sys, time, re, os, queue, threading, shutil, json
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.by import By
    from webdriver_manager.chrome import ChromeDriverManager
except ImportError:
    print("Eksik: pip install selenium openpyxl beautifulsoup4 webdriver-manager")
    sys.exit(1)

PARALEL_TARAYICI = 7
SAYFA_BEKLEME    = 5
MAX_DENEME       = 2
TXT_DOSYA        = "hisseisimleri.txt"
EXCEL_DOSYA      = "BISTTemelVeriler.xlsx"

cekilen_sayisi = 0
sayac_kilidi   = threading.Lock()

# Tüm sütunlar — (Excel başlığı, kayıt yolu)
OZET_SUTUNLAR = [
    # ── Kimlik ──────────────────────────────────────────
    ("Kod",                    "kod"),
    ("Şirket Adı",             "sirket_adi"),
    # ── Çarpanlar (Tablo 8) ─────────────────────────────
    ("F/K",                    "cari.F/K"),
    ("FD/FAVÖK",               "cari.FD/FAVÖK"),
    ("PD/DD",                  "cari.PD/DD"),
    ("FD/Satışlar",            "cari.FD/Satışlar"),
    ("Yabancı Oranı (%)",      "cari.Yabancı Oranı"),
    ("Piyasa Değeri (mnTL)",   "cari.Piyasa Değeri"),
    ("Net Borç (mnTL)",        "cari.Net Borç"),
    ("Halka Açıklık (%)",      "cari.Halka Açıklık"),
    # ── Performans (Tablo 9) ────────────────────────────
    ("1 Gün TL (%)",           "perf.1G_TL"),
    ("1 Hafta TL (%)",         "perf.1H_TL"),
    ("1 Ay TL (%)",            "perf.1A_TL"),
    ("Yıl İçi TL (%)",        "perf.YTD_TL"),
    ("1 Gün Göreceli (%)",     "perf.1G_Gor"),
    # ── Fiyat Hareketi (Tablo 4) ────────────────────────
    ("3A Min Fiyat (TL)",      "fiyat.min"),
    ("3A Max Fiyat (TL)",      "fiyat.max"),
    ("3A Fiyat Değişim (TL)", "fiyat.degisim"),
    # ── Endeks Üyeliği (Tablo 5) ────────────────────────
    ("XU100",                  "endeks.XU100"),
    ("XU050",                  "endeks.XU050"),
    ("XU030",                  "endeks.XU030"),
    # ── Mali Özet (Tablo 7) ─────────────────────────────
    ("Özkaynaklar (mnTL)",     "mali.Özkaynaklar"),
    ("Öd. Sermaye (mnTL)",     "mali.Ödenmiş Sermaye"),
    ("Net Kâr (mnTL)",         "mali.Net Kâr"),
    # ── Künye (Tablo 6) ─────────────────────────────────
    ("Ünvanı",                 "kunye.Ünvanı"),
    ("Kuruluş",                "kunye.Kuruluş"),
    ("Faal Alanı",             "kunye.Faal Alanı"),
    ("Telefon",                "kunye.Telefon"),
    ("Adres",                  "kunye.Adres"),
]

def temizle(x):
    return " ".join(str(x).split()).strip() if x else ""

def sayi(x):
    if x is None: return None
    s = str(x).strip().replace(".", "").replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    try: return float(s)
    except ValueError: return None

def deger_al(kayit, yol):
    parcalar = yol.split(".", 1)
    val = kayit.get(parcalar[0])
    if len(parcalar) == 1: return val
    if isinstance(val, dict): return deger_al(val, parcalar[1])
    return None

def oku_txt(dosya):
    if not os.path.exists(dosya):
        print(f"HATA: '{dosya}' bulunamadi!")
        return []
    with open(dosya, "r", encoding="utf-8") as f:
        kodlar = [l.strip().upper() for l in f if l.strip()]
    if not kodlar:
        print(f"HATA: '{dosya}' bos!")
    return kodlar

def chrome_ve_driver_bul():
    driver_path = shutil.which("chromedriver")
    if driver_path:
        print(f"  [chromedriver] PATH: {driver_path}")
    else:
        print("  [chromedriver] webdriver-manager ile indiriliyor...")
        driver_path = ChromeDriverManager().install()
    chrome_bin = (
        shutil.which("chrome") or shutil.which("google-chrome") or
        shutil.which("google-chrome-stable") or
        shutil.which("chromium-browser") or shutil.which("chromium")
    )
    if chrome_bin:
        print(f"  [chrome]       PATH: {chrome_bin}")
    return driver_path, chrome_bin

def chrome_olustur(driver_path, chrome_bin=None):
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    if chrome_bin:
        opts.binary_location = chrome_bin
    return webdriver.Chrome(service=Service(driver_path), options=opts)

def sayfa_cek(driver, kod):
    url = (
        "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/"
        f"sirket-karti.aspx?hisse={kod}"
    )
    for deneme in range(1, MAX_DENEME + 1):
        try:
            driver.get(url)
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table"))
            )
            time.sleep(SAYFA_BEKLEME)
            return driver.page_source
        except Exception as e:
            print(f"\n  [UYARI] {kod} deneme {deneme}/{MAX_DENEME}: {type(e).__name__}")
            if deneme < MAX_DENEME: time.sleep(3)
    try: return driver.page_source
    except: return ""

def ayristir(html, kod):
    soup = BeautifulSoup(html, "html.parser")
    kayit = {
        "kod": kod, "sirket_adi": kod,
        "kunye": {}, "cari": {}, "mali": {},
        "perf": {}, "fiyat": {}, "endeks": {},
    }

    # Şirket adı
    h1 = soup.find("h1")
    if h1:
        kayit["sirket_adi"] = temizle(h1.get_text()).split("|")[0].strip()

    tablolar = soup.find_all("table")

    for i, tablo in enumerate(tablolar):
        satirlar = tablo.find_all("tr")

        # ── Tablo 4: Fiyat Hareketi ─────────────────────
        if i == 4 and len(satirlar) >= 2:
            for satir in satirlar[1:]:
                hucreler = satir.find_all(["td","th"])
                if len(hucreler) >= 4:
                    baslik = temizle(hucreler[0].get_text())
                    if "Fiyat" in baslik:
                        kayit["fiyat"]["min"]      = sayi(hucreler[1].get_text())
                        kayit["fiyat"]["max"]      = sayi(hucreler[2].get_text())
                        kayit["fiyat"]["degisim"]  = sayi(hucreler[3].get_text())

        # ── Tablo 5: Endeks üyeliği ─────────────────────
        elif i == 5 and len(satirlar) >= 2:
            basliklar = [temizle(h.get_text()) for h in satirlar[0].find_all(["td","th"])]
            degerler  = [temizle(d.get_text()) for d in satirlar[1].find_all(["td","th"])]
            for b, d in zip(basliklar, degerler):
                if "XU100" in b: kayit["endeks"]["XU100"] = d
                elif "XU050" in b: kayit["endeks"]["XU050"] = d
                elif "XU030" in b: kayit["endeks"]["XU030"] = d

        # ── Tablo 6: Künye ──────────────────────────────
        elif i == 6:
            for satir in satirlar:
                h = satir.find_all(["td","th"])
                if len(h) >= 2:
                    k = temizle(h[0].get_text())
                    v = temizle(h[1].get_text())
                    if k in ("Ünvanı","Kuruluş","Faal Alanı","Telefon","Faks","Adres"):
                        kayit["kunye"][k] = v

        # ── Tablo 7: Mali özet ──────────────────────────
        elif i == 7:
            for satir in satirlar:
                h = satir.find_all(["td","th"])
                if len(h) >= 2:
                    k = temizle(h[0].get_text())
                    v = temizle(h[1].get_text())
                    if k == "Özkaynaklar":        kayit["mali"]["Özkaynaklar"]      = sayi(v)
                    elif k == "Ödenmiş Sermaye":  kayit["mali"]["Ödenmiş Sermaye"]  = sayi(v)
                    elif k == "Net Kâr":          kayit["mali"]["Net Kâr"]          = sayi(v)

        # ── Tablo 8: Çarpanlar ──────────────────────────
        elif i == 8:
            for satir in satirlar:
                h = satir.find_all(["td","th"])
                if len(h) >= 2:
                    k = temizle(h[0].get_text())
                    v = temizle(h[1].get_text())
                    if   k == "F/K":                    kayit["cari"]["F/K"]           = sayi(v)
                    elif k == "FD/FAVÖK":               kayit["cari"]["FD/FAVÖK"]      = sayi(v)
                    elif k == "PD/DD":                  kayit["cari"]["PD/DD"]         = sayi(v)
                    elif k == "FD/Satışlar":            kayit["cari"]["FD/Satışlar"]   = sayi(v)
                    elif k == "Yabancı Oranı (%)":      kayit["cari"]["Yabancı Oranı"] = sayi(v)
                    elif k == "Piyasa Değeri":          kayit["cari"]["Piyasa Değeri"] = sayi(v)
                    elif k == "Net Borç":               kayit["cari"]["Net Borç"]      = sayi(v)
                    elif k == "Halka Açıklık Oranı (%)": kayit["cari"]["Halka Açıklık"] = sayi(v)

        # ── Tablo 9: Performans ─────────────────────────
        elif i == 9 and len(satirlar) >= 4:
            # Satır 1=TL, Satır 3=Göreceli
            tl  = satirlar[1].find_all(["td","th"])
            gor = satirlar[3].find_all(["td","th"])
            if len(tl) >= 5:
                kayit["perf"]["1G_TL"]  = sayi(tl[1].get_text())
                kayit["perf"]["1H_TL"]  = sayi(tl[2].get_text())
                kayit["perf"]["1A_TL"]  = sayi(tl[3].get_text())
                kayit["perf"]["YTD_TL"] = sayi(tl[4].get_text())
            if len(gor) >= 2:
                kayit["perf"]["1G_Gor"] = sayi(gor[1].get_text())

    if not kayit["cari"]:
        snippet = soup.get_text()[:200].replace("\n", " ")
        print(f"\n  [DEBUG] {kod} veri YOK. Sayfa: {snippet}")

    return kayit

def worker_calis(kodlar_q, sonuclar, toplam, worker_id, driver_path, chrome_bin):
    global cekilen_sayisi
    driver = None
    try:
        driver = chrome_olustur(driver_path, chrome_bin)
    except Exception as e:
        print(f"\n  [HATA] Worker-{worker_id} Chrome acilmadi: {e}")
        while not kodlar_q.empty():
            try:
                sira, kod = kodlar_q.get_nowait()
                sonuclar[sira] = {"kod": kod, "sirket_adi": "Chrome hatasi"}
                kodlar_q.task_done()
            except queue.Empty: break
        return

    try:
        while True:
            try: sira, kod = kodlar_q.get_nowait()
            except queue.Empty: break
            try:
                html  = sayfa_cek(driver, kod)
                kayit = ayristir(html, kod)
                sonuclar[sira] = kayit
            except Exception as e:
                print(f"\n  [HATA] Worker-{worker_id} | {kod}: {e}")
                sonuclar[sira] = {"kod": kod, "sirket_adi": "HATA"}
            with sayac_kilidi:
                cekilen_sayisi += 1
                yuzde = (cekilen_sayisi / toplam) * 100
                sys.stdout.write(
                    f"\r  >> %{yuzde:.1f} | {cekilen_sayisi}/{toplam} ({kod}){'':10}"
                )
                sys.stdout.flush()
            kodlar_q.task_done()
    finally:
        if driver:
            try: driver.quit()
            except: pass

def excel_yaz(veri_listesi, dosya_adi):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hisse_Verileri"

    # Başlık stili
    baslik_fill = PatternFill("solid", fgColor="1F4E79")
    baslik_font = Font(bold=True, color="FFFFFF", size=10)

    for ci, (baslik, _) in enumerate(OZET_SUTUNLAR, 1):
        hucre = ws.cell(1, ci, baslik)
        hucre.font = baslik_font
        hucre.fill = baslik_fill
        hucre.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    dolu = bos = 0
    for ri, kayit in enumerate(veri_listesi, 2):
        if not kayit:
            bos += 1
            continue
        if not kayit.get("cari"):
            bos += 1
            ws.cell(ri, 1, kayit.get("kod",""))
            ws.cell(ri, 2, kayit.get("sirket_adi",""))
            continue
        dolu += 1
        for ci, (_, yol) in enumerate(OZET_SUTUNLAR, 1):
            ws.cell(ri, ci, deger_al(kayit, yol))

    # Kolon genişlikleri
    genislikler = {
        1:8, 2:25, 3:7, 4:9, 5:7, 6:10, 7:12, 8:15, 9:14, 10:13,
        11:11, 12:12, 13:10, 14:11, 15:13,
        16:13, 17:13, 18:15,
        19:12, 20:12, 21:12,
        22:14, 23:14, 24:14,
        25:20, 26:12, 27:40, 28:14, 29:40,
    }
    for ci, genislik in genislikler.items():
        ws.column_dimensions[get_column_letter(ci)].width = genislik

    # Dondur (başlık satırı)
    ws.freeze_panes = "A2"

    wb.save(dosya_adi)
    print(f"\n  Excel: {dosya_adi} — {dolu} hisse / {bos} bos")

def json_yaz(veri_listesi, dosya_adi):
    """
    veri_listesi: excel_yaz ile aynı formatta kayıt listesi
    dosya_adi: çıktı JSON dosyasının adı (örn. BISTTemelVeriler.json)
    """
    tum_veriler = []
    for kayit in veri_listesi:
        if not kayit:
            # Boş kayıtları atla veya None ile geç
            tum_veriler.append(None)
            continue
        satir = {}
        for baslik, yol in OZET_SUTUNLAR:
            deger = deger_al(kayit, yol)
            # JSON'da None, float, str olarak kalsın
            satir[baslik] = deger
        tum_veriler.append(satir)

    with open(dosya_adi, "w", encoding="utf-8") as f:
        json.dump(tum_veriler, f, ensure_ascii=False, indent=2)

    print(f"  JSON: {dosya_adi} — {len([v for v in tum_veriler if v])} hisse kaydedildi.")

def main():
    kodlar = oku_txt(TXT_DOSYA)
    if not kodlar: return
    toplam = len(kodlar)
    print(f"\n{toplam} hisse isleniyor ({PARALEL_TARAYICI} paralel tarayici)\n")
    driver_path, chrome_bin = chrome_ve_driver_bul()
    print()
    kodlar_q = queue.Queue()
    for i, k in enumerate(kodlar): kodlar_q.put((i, k))
    sonuclar = [None] * toplam
    threads = []
    for wid in range(1, min(PARALEL_TARAYICI, toplam) + 1):
        t = threading.Thread(
            target=worker_calis,
            args=(kodlar_q, sonuclar, toplam, wid, driver_path, chrome_bin),
            daemon=True,
        )
        t.start()
        threads.append(t)
    for t in threads: t.join()
    excel_yaz(sonuclar, EXCEL_DOSYA)
    json_yaz(sonuclar, "BISTTemelVeriler.json")
    print(f"Tamamlandi -> {EXCEL_DOSYA}")

if __name__ == "__main__":
    main()
